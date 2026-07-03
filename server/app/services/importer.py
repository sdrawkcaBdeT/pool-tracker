"""One-shot import of the historical pool.xlsx game log.

Every transformation is deterministic and reported: alias merges, the
fall-2017 balls-left column swap, cutthroat finish-place recovery, doubles
participants, and skipped junk rows. Nothing is silently altered — original
values land in the game's notes whenever a correction is applied.

Conventions established while profiling the workbook (2026-07-02):
- The data sheet is whichever sheet has a 'GameCount' header in row 1.
- Real game rows are the ones whose date cell is a datetime. That excludes
  two empty stub rows (GameCount 611/612) and the summary block at the
  bottom ('games/wins/losses/winrate').
- win/loss is from the owner's (ted's) perspective.
- Games 91-134 include ten rows whose loser/winner balls-left values are
  impossible as recorded (regulation or scratch-on-8 with the wrong column
  populated) and become rule-legal when swapped; they cluster Sep 4 -
  Oct 28, 2017. Diagnosis: the two columns were entered in reverse order
  during that stretch. The importer swaps them and says so in notes.
- Games 321 and 585 are inconsistent but not fixable by the swap; they
  import as recorded and get flagged.
- Old-style cutthroat rows store the first-out player's name in 'loser
  remainder' and the winner's name in 'winner remaining' -> finish places
  (winner=1, first-out=3, remaining participant=2). Two 2020 rows recorded
  ball counts instead; those keep the counts in notes, finish order unknown
  beyond the recorded win/loss.
- The owner's own player row is the player whose name equals the user's
  username ('ted'); stats rely on this convention.
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as dt
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import REPO_ROOT
from app.database import create_tables, get_sessionmaker
from app.models import Game, GamePlayer, Player, PlaySession, User, Venue

OWNER_USERNAME = "ted"
OWNER_DISPLAY_NAME = "Ted Janka"

VENUE_ALIASES = {
    "adams": "adams house",
}
PLAYER_ALIASES = {
    "christain": "christian",
}
# Multi-opponent cells written without a comma separator.
SPACE_SEPARATED_OPPONENTS = {"adam josh", "josh jered"}

WIN_TYPE_MAP = {
    "regulation": "regulation",
    "early 8": "early_8",
    "scratch on 8": "scratch_on_8",
    "wrong pocket": "wrong_pocket",
    "win on break": "win_on_break",
}

# Fall-2017 window where the two balls-left columns were entered in reverse
# order. Each of these rows is impossible as recorded and uniquely valid
# when swapped.
SWAP_SEQS = {91, 117, 118, 121, 124, 126, 128, 132, 133, 134}
# Inconsistent as recorded, not fixable by the swap; imported as-is.
FLAG_SEQS = {321, 585}


def norm(value) -> str:
    return " ".join(str(value).strip().split()).lower()


def canon_venue(value) -> str:
    name = norm(value)
    return VENUE_ALIASES.get(name, name)


def canon_player(value) -> str:
    name = norm(value)
    return PLAYER_ALIASES.get(name, name)


def split_opponents(value) -> list[str]:
    name = norm(value)
    if name in SPACE_SEPARATED_OPPONENTS:
        return [canon_player(part) for part in name.split()]
    if "," in name:
        return [canon_player(part) for part in name.split(",")]
    return [canon_player(name)]


@dataclass
class ImportReport:
    games: int = 0
    sessions: int = 0
    venues: int = 0
    players: int = 0
    wins: int = 0
    losses: int = 0
    merges: list[str] = field(default_factory=list)
    corrections: list[str] = field(default_factory=list)
    flags: list[str] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)

    def print(self) -> None:
        print("=== import report ===")
        print(
            f"games={self.games} ({self.wins}W-{self.losses}L)  "
            f"sessions={self.sessions}  venues={self.venues}  players={self.players}"
        )
        for title, lines in (
            ("merges", sorted(set(self.merges))),
            ("corrections", self.corrections),
            ("flags", self.flags),
            ("skipped", self.skipped),
        ):
            print(f"--- {title} ({len(lines)}) ---")
            for line in lines:
                print(f"  {line}")


def find_data_sheet(workbook) -> object:
    for sheet in workbook.worksheets:
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if any(str(cell).strip() == "GameCount" for cell in header if cell is not None):
            return sheet
    raise ValueError("No sheet with a 'GameCount' header found")


@dataclass
class ParsedGame:
    seq: int
    date: dt.date
    venue: str
    game_type: str
    result: str
    win_type: str | None
    breaker: str
    loser_balls_left: int | None
    winner_balls_left: int | None
    opponents: list[str]
    teammates: list[str]
    finish_places: dict[str, int]
    notes: str


def parse_workbook(xlsx_path: Path, report: ImportReport) -> list[ParsedGame]:
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = find_data_sheet(workbook)
    parsed: list[ParsedGame] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        if not isinstance(row[0], dt.datetime):
            report.skipped.append(f"non-game row: {[str(v) for v in row[:5] if v is not None]}")
            continue

        seq = int(row[1])
        date = row[0].date()
        raw_venue, raw_opp, raw_win_type, raw_break = row[2], row[3], row[5], row[7]
        venue = canon_venue(raw_venue)
        if venue != str(raw_venue):
            report.merges.append(f"venue {str(raw_venue)!r} -> {venue!r}")

        opponents = split_opponents(raw_opp)
        if ",".join(opponents) != str(raw_opp):
            report.merges.append(f"opponent {str(raw_opp)!r} -> {opponents!r}")

        breaker = canon_player(raw_break)
        if breaker != str(raw_break):
            report.merges.append(f"breaker {str(raw_break)!r} -> {breaker!r}")

        result = "win" if row[4] == 1 else "loss"
        win_type_raw = norm(raw_win_type)
        notes_parts = [norm(row[9])] if row[9] is not None else []

        loser_balls, winner_balls = row[6], row[8]
        teammates: list[str] = []
        finish_places: dict[str, int] = {}

        if win_type_raw == "cutthroat":
            game_type = "cutthroat"
            win_type = None
            if isinstance(loser_balls, str):
                first_out = canon_player(loser_balls)
                winner = canon_player(winner_balls)
                participants = [OWNER_USERNAME, *opponents]
                middle = [p for p in participants if p not in (first_out, winner)]
                finish_places = {winner: 1, first_out: 3}
                if len(middle) == 1:
                    finish_places[middle[0]] = 2
                report.corrections.append(
                    f"game {seq}: cutthroat finish order recovered "
                    f"(1st {winner}, 3rd {first_out})"
                )
            else:
                notes_parts.append(
                    f"import: cutthroat recorded ball counts loser={loser_balls} "
                    f"winner={winner_balls}; finish order unknown"
                )
                if result == "win":
                    finish_places = {OWNER_USERNAME: 1}
                report.flags.append(f"game {seq}: cutthroat with ball counts, finish order unknown")
            loser_balls = None
            winner_balls = None
        else:
            game_type = "doubles" if len(opponents) > 1 else "singles"
            win_type = WIN_TYPE_MAP[win_type_raw]
            if seq in SWAP_SEQS:
                notes_parts.append(
                    f"import: swapped balls-left columns (fall-2017 entry swap); "
                    f"recorded loser={loser_balls} winner={winner_balls}"
                )
                loser_balls, winner_balls = winner_balls, loser_balls
                report.corrections.append(
                    f"game {seq}: balls-left columns swapped "
                    f"(now loser={loser_balls} winner={winner_balls})"
                )
            if seq in FLAG_SEQS:
                notes_parts.append(
                    "import: balls-left inconsistent with scratch-on-8 as recorded"
                )
                report.flags.append(
                    f"game {seq}: {win_type} with loser_balls_left={loser_balls} "
                    f"winner_balls_left={winner_balls}, imported as recorded"
                )
            if game_type == "doubles":
                note = notes_parts[0] if notes_parts else ""
                if note.endswith(("team", "teammate")):
                    teammate = canon_player(note.removesuffix("teammate").removesuffix("team"))
                    teammates = [teammate]
                    report.corrections.append(
                        f"game {seq}: teammate {teammate} recovered from note {note!r}"
                    )
                else:
                    report.flags.append(f"game {seq}: doubles vs {opponents}, teammate unknown")

        parsed.append(
            ParsedGame(
                seq=seq,
                date=date,
                venue=venue,
                game_type=game_type,
                result=result,
                win_type=win_type,
                breaker=breaker,
                loser_balls_left=loser_balls if isinstance(loser_balls, int) else None,
                winner_balls_left=winner_balls if isinstance(winner_balls, int) else None,
                opponents=opponents,
                teammates=teammates,
                finish_places=finish_places,
                notes="; ".join(part for part in notes_parts if part),
            )
        )

    seqs = [game.seq for game in parsed]
    if sorted(seqs) != list(range(1, len(parsed) + 1)):
        raise ValueError("GameCount sequence is not contiguous 1..N after filtering")
    return parsed


async def run_import(db: AsyncSession, parsed: list[ParsedGame], report: ImportReport, replace: bool) -> None:
    user = (
        await db.execute(select(User).where(User.username == OWNER_USERNAME))
    ).scalar_one_or_none()
    if user is None:
        user = User(username=OWNER_USERNAME, display_name=OWNER_DISPLAY_NAME)
        db.add(user)
        await db.flush()

    existing = (
        await db.execute(select(Game.id).where(Game.user_id == user.id).limit(1))
    ).scalar_one_or_none()
    if existing is not None:
        if not replace:
            raise SystemExit("Games already imported for this user; rerun with --replace")
        game_ids = select(Game.id).where(Game.user_id == user.id)
        await db.execute(delete(GamePlayer).where(GamePlayer.game_id.in_(game_ids)))
        await db.execute(delete(Game).where(Game.user_id == user.id))
        await db.execute(delete(PlaySession).where(PlaySession.user_id == user.id))
        await db.execute(delete(Player).where(Player.user_id == user.id))
        await db.execute(delete(Venue).where(Venue.user_id == user.id))

    players: dict[str, Player] = {}
    venues: dict[str, Venue] = {}
    sessions: dict[tuple[dt.date, str], PlaySession] = {}

    async def get_player(name: str) -> Player:
        if name not in players:
            players[name] = Player(user_id=user.id, name=name)
            db.add(players[name])
            await db.flush()
        return players[name]

    await get_player(OWNER_USERNAME)

    for game in parsed:
        if game.venue not in venues:
            venues[game.venue] = Venue(user_id=user.id, name=game.venue)
            db.add(venues[game.venue])
            await db.flush()
        session_key = (game.date, game.venue)
        if session_key not in sessions:
            sessions[session_key] = PlaySession(
                user_id=user.id, date=game.date, venue_id=venues[game.venue].id
            )
            db.add(sessions[session_key])
            await db.flush()

        row = Game(
            user_id=user.id,
            session_id=sessions[session_key].id,
            seq=game.seq,
            game_type=game.game_type,
            result=game.result,
            win_type=game.win_type,
            breaker_player_id=(await get_player(game.breaker)).id,
            loser_balls_left=game.loser_balls_left,
            winner_balls_left=game.winner_balls_left,
            notes=game.notes,
            entry_mode="import",
        )
        db.add(row)
        await db.flush()

        db.add(
            GamePlayer(
                game_id=row.id,
                player_id=players[OWNER_USERNAME].id,
                side="me",
                finish_place=game.finish_places.get(OWNER_USERNAME),
            )
        )
        for name in game.opponents:
            db.add(
                GamePlayer(
                    game_id=row.id,
                    player_id=(await get_player(name)).id,
                    side="opponent",
                    finish_place=game.finish_places.get(name),
                )
            )
        for name in game.teammates:
            db.add(
                GamePlayer(
                    game_id=row.id,
                    player_id=(await get_player(name)).id,
                    side="teammate",
                    finish_place=game.finish_places.get(name),
                )
            )

        report.games += 1
        if game.result == "win":
            report.wins += 1
        else:
            report.losses += 1

    report.sessions = len(sessions)
    report.venues = len(venues)
    report.players = len(players)
    await db.commit()


async def import_if_empty() -> None:
    """Startup hook (AUTO_IMPORT=true): seed the DB from the bundled workbook
    when the owner has no games yet. A populated DB is left untouched."""
    from app.config import get_settings

    xlsx_path = Path(get_settings().data_dir) / "pool.xlsx"
    if not xlsx_path.exists():
        print(f"auto-import: {xlsx_path} not found, skipping")
        return
    async with get_sessionmaker()() as db:
        user = (
            await db.execute(select(User).where(User.username == OWNER_USERNAME))
        ).scalar_one_or_none()
        if user is not None:
            existing = (
                await db.execute(select(Game.id).where(Game.user_id == user.id).limit(1))
            ).scalar_one_or_none()
            if existing is not None:
                return
    report = ImportReport()
    parsed = parse_workbook(xlsx_path, report)
    async with get_sessionmaker()() as db:
        await run_import(db, parsed, report, replace=False)
    report.print()


async def main_async(xlsx_path: Path, replace: bool) -> None:
    report = ImportReport()
    parsed = parse_workbook(xlsx_path, report)
    await create_tables()
    async with get_sessionmaker()() as db:
        await run_import(db, parsed, report, replace)
    report.print()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=REPO_ROOT / "data" / "raw" / "pool.xlsx",
        help="Path to the historical workbook",
    )
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Delete previously imported data for the owner before importing",
    )
    args = parser.parse_args()
    asyncio.run(main_async(args.xlsx, args.replace))


if __name__ == "__main__":
    main()
